class FileParse():
    def __init__(self,filePath):

        self.parse(filePath)

    def parse(self,filePath):

        def getPf(lineList):
            pf=lineList[1]
            pf= pf[6:8]
            return pf

        def getPs(lineList):
            ps=lineList[2]
            ps=ps[6:8]
            return ps

        def getSa(lineList):
            sa=lineList[3]
            if sa[0:2]=="sa":
                sa =sa[4:8]
                return sa
            else:
                return ""
            

        def getPgLabel(lineList):
            i =0
            label =lineList[3]
            if label[0:2]=="sa":
                label = lineList[4]
                label = label[7:len(label)]
                while i != (len(lineList)-5):
                    i+=1
                    label=label+" "+lineList[4+i]
                label = label[:-2]
                
                return label
            else:
                label = label[7:len(label)]
                while i != (len(lineList)-4):
                    i+=1
                    label=label+" "+lineList[3+i]
                label = label[:-2]
                return label

        def getByte(lineList):
            byte=lineList[1]
            byte =byte[-2:-1]
            return byte

        def getBit(lineList):
            bit = lineList[2]
            bit =bit[-2:-1]
            return bit

        def getLength(lineList):

            if lineList[3][9].isdigit():
                length= lineList[3]
                length=length[8:10]
                return length
            else:
                return lineList[3][8]

        def getSelectfieldLabel(lineList, hasIDfield):
            
            fieldLabel= lineList[4]
            i=0
            fieldLabel = fieldLabel[7:len(fieldLabel)]
            while i != (len(lineList)-5):
                i+=1
                fieldLabel=fieldLabel+" "+lineList[4+i]
            fieldLabel = fieldLabel[:-2]
            if hasIDfield==False and fieldLabel.count("spn")==0:
                return fieldLabel
            else:
                fieldLabel=fieldLabel[:fieldLabel.rfind('\"')]
                fieldLabel=fieldLabel[:fieldLabel.rfind('\"')]
                labelEndIndex=fieldLabel.rfind('\"')
                #fieldLabel=fieldLabel[:labelEndIndex]
                
                fieldLabel=fieldLabel[(fieldLabel.rfind('\"')+1):]
                #fieldLabel=fieldLabel[(fieldLabel.rfind('\"')+1):labelEndIndex]
                return fieldLabel
            

        def getSelectfieldID(lineList):
            id=lineList[-1]
            id=id[(id.find('\"')+1):id.rfind('\"')]
            return id
            

        def getTextfieldLabel(lineList):
            textfieldLabel=lineList[5]
            i=0
            textfieldLabel = textfieldLabel[7:len(textfieldLabel)]
            while i != (len(lineList)-6):
                i+=1
                textfieldLabel=textfieldLabel+" "+lineList[5+i]
            textfieldLabel = textfieldLabel[:-4]
            return textfieldLabel
            

        def getBase(lineList):
            base =lineList[1]
            base= base[6:-1]
            return base

        def getTextFieldType(lineList):
            textFieldType=lineList[4]
            textFieldType=textFieldType[6:-1]
            return textFieldType


        def getSpn(lineList):
            if lineList[-1].count("spn")==1:
                spn=lineList[-1]
            
                spn=spn[(spn.find('\"')+1):spn.rfind('\"')]
                return spn
            else:
                return ""

        from openpyxl import Workbook
        from openpyxl.styles import Alignment



        wb = Workbook()
        ws = wb.active # grab the active worksheet

        ali=Alignment(horizontal='general',vertical='bottom',text_rotation=0,wrap_text=True,shrink_to_fit=False,indent=0)

        ################## set up the header########
        ws['A1']="Priority"
        ws['B1']="PGN"
        ws['C1']="SA"
        ws['D1']="ID"
        ws['E1']="PGN abreviation"
        ws['F1']= "PGN name"
        ws['G1']="PGN description"
        ws['H1']="Data length"
        ws['I1']="Parameter name"
        ws['J1']="Parameter description"
        ws['K1']="Start position"
        ws['L1']="Bit lenght"
        ws['M1']="Multiplexor start positon"
        ws['N1']="Multiplexor length"
        ws['O1']="Multiplexor value"
        ws['P1']="SPN"
        ws['Q1']="Factor"
        ws['R1']="Offset"
        ws['S1']="Unit"
        ws['T1']="Minimum"
        ws['U1']="Maximum"
        ws['V1']="base"
        #############################################


        inTag = False
        inJref = False
        hasIDfield=False
        row =3
        fileHandle=open(filePath,'r')
        
        #savePath=filePath.split('/')
        #savePath=savePath[-1]
        #savePath=savePath.split('.')
        
        #savePath=savePath[0]+"_jsyref.jsym"
        
        #jsymRef = open(savePath[0]+"_jsyref.jsym", 'w+')
        #savePath=filePath[0:(filePath.rfind("\\")+1)]
        jsymRef = open("jsymref.jsym", 'w+')

        print("PGN    SA    PGN Name  Data length      Parameter name         Start positoin   Bit length      base")
        
        for line in fileHandle:
            lineList= line.split()
            if lineList==[]:
                continue
            
            if lineList[0] == "<?xml":
                inJref=True

            if lineList[0] == "<pg":
                inJref=False

                
            if lineList[0] == "<pg":
                inTag=True
                pf =getPf(lineList)
                ps=getPs(lineList)
                sa=getSa(lineList)
                dataLength="8"
                pgLabel=getPgLabel(lineList)

            elif inJref:
                jsymRef.write(line)


            elif lineList[0]=="<idfield":
                hasIDfield=True
                multiplexerByte=getByte(lineList)
                multiplexerBit=getBit(lineList)
                multiplexerLength=getLength(lineList)
                #ws["M"+str(row)]=multiplexerByte+"."+multiplexerBit
                #ws["N"+str(row)]=multiplexerLength

                
            elif lineList[0] =="<selectfield":
                byte=getByte(lineList)
                bit=getBit(lineList)
                length=getLength(lineList)
                fieldLabel=getSelectfieldLabel(lineList,hasIDfield)
                spn=getSpn(lineList)
                if hasIDfield:
                    ID=getSelectfieldID(lineList)
                
                if inTag:
                    ws["B"+str(row)]="0x"+pf+ps
                    ws["C"+str(row)]=sa
                    ws["F"+str(row)]=pgLabel
                    ws["H"+str(row)]="8"
                    ws["I"+str(row)]=fieldLabel
                    ws["K"+str(row)]=byte+"."+bit
                    ws["L"+str(row)]=length
                    #ws["O"+str(row)]=ID
                    ws["P"+str(row)]=spn

                    if hasIDfield:
                        ws["M"+str(row)]=multiplexerByte+"."+multiplexerBit
                        ws["N"+str(row)]=multiplexerLength
                        ws["O"+str(row)]=ID
                        
                    
                    #print("0x"+pf+ps+" "+sa+"  "+pgLabel+"     "+"8"+"       "+fieldLabel, end='')
                    #print(byte.rjust(70-len("0x"+pf+ps+" "+sa+"  "+pgLabel+"     "+"8"+"       "+fieldLabel))+"."+bit+"          "+length+"  ", end='         ')

                    
                    
            elif  lineList[0] =="<textfield":
                byte=getByte(lineList)
                bit=getBit(lineList)
                length=getLength(lineList)
                textFieldType=getTextFieldType(lineList)
                textfieldLabel=getTextfieldLabel(lineList)
                
                
                ws["B"+str(row)]="0x"+pf+ps
                ws["C"+str(row)]=sa
                ws["F"+str(row)]=pgLabel
                ws["H"+str(row)]="8"
                ws["I"+str(row)]=textfieldLabel
                ws["K"+str(row)]=byte+"."+bit
                ws["L"+str(row)]=length
                ws["V"+str(row)]=textFieldType
                
                #print("0x"+pf+ps+" "+sa+"  "+pgLabel+"       "+"8"+"       "+textfieldLabel, end='')
                #print(byte.rjust(70-len("0x"+pf+ps+" "+sa+"  "+pgLabel+"       "+"8"+"       "+textfieldLabel))+"."+bit+"          "+length+"           "+textFieldType)
                row = row+1

            elif lineList[0]=="<select" and inTag:
                if inTag:
                    base=getBase(lineList)
                    ws["V"+str(row)]=base
                    row=row+1

            elif lineList[0]=="<select>":
                if inTag:
                    ws["V"+str(row)]=""
                    ws["V"+str(row)].style='Bad'
                    row=row+1
                    
                    

            elif lineList[0]=="</pg>":
                inTag=False
                hasIDfield=False
                row = row+3
                #print()
                #print()
                #print()

            
                
            
            
        jsymRef.write("</j1939system>") 
        jsymRef.close() 
        fileHandle.close()

        wb.save("test.xlsx")
        ##wb.save(savePath[0]+".xlsx")
    


    
